"""The terminology contract, plus parsing shared by the Swedish classification files.

Adding a terminology system must mean "write one loader", not "restructure the
project" -- so everything downstream (retrieval, reranking, proposals) speaks
`Concept` and nothing else.

FILE FORMAT PROVENANCE
----------------------
The Swedish classification code-text files are published as tab-separated text
files. The structure encoded here is taken from the publisher's own file
description PDFs (see LICENSING.md for the documents and the date checked):

  * UTF-8, tab-separated, first row holds column headers.
  * Every cell value is surrounded by double quotes.
  * "En kod upptar en eller flera rader beroende pa antalet egenskaper" --
    one code spans one or more rows, one row per repeated property value. Rows
    for the same code must therefore be merged, which is what
    `read_classification_tsv` does.
  * ICD-10-SE additionally marks several columns as RICH_TEXT holding HTML.

The exact header spellings are matched tolerantly (case, whitespace and dash
folding, plus aliases) because the description PDFs render the column names in
a wrapped table; see PHASE1_REPORT.md, marked ASSUMED.
"""

from __future__ import annotations

import csv
import html
import logging
import re
import unicodedata
from collections.abc import Iterable, Iterator, Mapping, Sequence
from pathlib import Path
from typing import Literal, Protocol, runtime_checkable

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

SystemId = Literal["icd10se", "kva", "snomed"]


class TerminologyError(RuntimeError):
    """Base class for terminology-layer failures."""


class TerminologyLicenceRequired(TerminologyError):
    """Raised when content cannot be shipped or loaded for licensing reasons."""


class TerminologyFormatError(TerminologyError):
    """Raised when an input file does not match the documented official format."""


class Concept(BaseModel):
    system: SystemId
    version: str
    code: str
    preferred_term: str
    synonyms: list[str] = Field(default_factory=list)
    parent_code: str | None = None
    is_leaf: bool
    chapter: str | None = None
    # The publisher's Beskrivning: prose explaining what the code covers. Kept
    # apart from `synonyms` because it is not a name for the concept -- it is
    # indexed at a lower weight, and never used as a display term.
    description: str = ""


@runtime_checkable
class TerminologySystem(Protocol):
    system_id: str

    def load(self, path: Path, version: str) -> Iterable[Concept]: ...

    def validate_code_format(self, code: str) -> bool: ...


# --------------------------------------------------------------------------- #
# Code intervals
# --------------------------------------------------------------------------- #

_INTERVAL_RE = re.compile(r"^[A-ZÅÄÖ0-9.]+-[A-ZÅÄÖ0-9.]+$", re.IGNORECASE)


def is_code_interval(code: str) -> bool:
    """True for chapter/section rows such as `I20-I25`.

    The publisher documents the Kod column as "en unik kod eller ett
    kodintervall". An interval names a group, not an assignable code, so it is
    loaded (the hierarchy needs it) but must never be proposed as a mapping.
    No valid ICD-10-SE or KVA code contains a hyphen, which is what makes this
    test safe.
    """
    return bool(_INTERVAL_RE.match(code.strip()))


# --------------------------------------------------------------------------- #
# TSV parsing
# --------------------------------------------------------------------------- #

_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")
_DASHES = str.maketrans({"‐": "-", "‑": "-", "‒": "-", "–": "-", "—": "-", "−": "-"})

# Canonical field name -> accepted header spellings (already folded).
_HEADER_ALIASES: Mapping[str, tuple[str, ...]] = {
    "code": ("kod",),
    "valid_from": ("giltig fran",),
    "parent": ("overordnad kod",),
    "title": ("titel",),
    "latin": ("latin",),
    "description": ("beskrivning",),
    "example": ("exempel",),
    "includes": ("innefattar",),
    "excludes": ("utesluter",),
    "note": ("anmarkning",),
    "coding_info": ("kodningsinformation",),
    "contents": ("innehall",),
    "abbreviations": ("forkortning(ar)", "forkortningar", "forkortning"),
    "manifestation": ("manifestation (*)/etiologi (t)", "manifestation/etiologi"),
    "manifestation_link": (
        "koppling manifestation (*)/etiologi (t)",
        "koppling manifestation/etiologi",
    ),
    "not_primary": ("ej huvuddiagnos",),
    "code_level": ("kodniva - kodspecifikation", "kodniva", "kodspecifikation"),
    "related_icf": ("relaterad icf-kod",),
    "related_icd10se": ("relaterad icd-10-se-kod",),
}

_ALIAS_TO_FIELD: dict[str, str] = {
    alias: field for field, aliases in _HEADER_ALIASES.items() for alias in aliases
}


def fold_header(raw: str) -> str:
    """Fold a column header to a comparison key.

    Case, surrounding whitespace, internal whitespace runs, dash variants and
    the Swedish diacritics are all normalised away, so `Kodniva - kodspecifikation`,
    `Kodnivå – kodspecifikation` and `KODNIVÅ  –  KODSPECIFIKATION` all agree.
    The dagger character used in the ICD-10-SE manifestation columns is folded
    to `t` for the same reason.
    """
    text = unicodedata.normalize("NFC", raw).strip().strip('"').translate(_DASHES)
    text = text.casefold()
    text = text.replace("Ɨ", "t").replace("ɨ", "t").replace("†", "t")
    text = (
        text.replace("å", "a")
        .replace("ä", "a")
        .replace("ö", "o")
        .replace("é", "e")
        .replace("ü", "u")
    )
    return _WS_RE.sub(" ", text).strip()


def strip_rich_text(value: str) -> str:
    """Turn a RICH_TEXT (HTML) cell into plain text.

    ICD-10-SE marks Beskrivning, Exempel, Innefattar, Utesluter, Anmarkning and
    Innehall as RICH_TEXT. Applied defensively to every cell: the KVA files
    declare plain TEXT, and stripping a string with no markup is a no-op.
    """
    text = _TAG_RE.sub(" ", value)
    text = html.unescape(text)
    return _WS_RE.sub(" ", text).strip()


def _map_header(header: Sequence[object], path: Path, required: Sequence[str]) -> list[str | None]:
    """Map a header row to canonical field names, one entry per column.

    Columns with no known alias map to None and are ignored, so an added column
    in a future release does not break a load. A missing *required* column is a
    hard error naming what was seen.
    """
    fields: list[str | None] = []
    for cell in header:
        text = "" if cell is None else str(cell)
        fields.append(_ALIAS_TO_FIELD.get(fold_header(text)) if text.strip() else None)

    missing = [name for name in required if name not in fields]
    if missing:
        seen = [("" if cell is None else str(cell)) for cell in header]
        raise TerminologyFormatError(
            f"{path}: missing required column(s) {missing}. "
            f"Saw headers {seen!r}. Expected an official code-text file "
            f"(see LICENSING.md)."
        )
    return fields


def _group_rows(
    fields: Sequence[str | None],
    rows: Iterable[tuple[int, Sequence[object]]],
    path: Path,
) -> dict[str, list[dict[str, str]]]:
    """Collect cells into one entry per code, preserving first-seen order.

    A single code spans several rows when it carries several repeated
    properties, so rows sharing a Kod value are merged rather than treated as
    separate concepts.
    """
    groups: dict[str, list[dict[str, str]]] = {}
    for line_no, raw_row in rows:
        if not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue
        row: dict[str, str] = {}
        for field, cell in zip(fields, raw_row, strict=False):
            if field is None or cell is None:
                continue
            value = strip_rich_text(str(cell))
            if value:
                row[field] = value
        code = row.get("code", "").strip()
        if not code:
            raise TerminologyFormatError(f"{path}:{line_no}: row has no Kod value")
        groups.setdefault(code, []).append(row)
    return groups


def read_classification_file(
    path: Path, *, required: Sequence[str] = ("code", "title")
) -> Iterator[tuple[str, list[dict[str, str]]]]:
    """Read an official code-text file, whichever format it is published in.

    The publishers distribute these as tab-separated text and as spreadsheets;
    which one a given classification offers has changed over time and differs
    between classifications. Dispatching on the extension means a loader is
    written once and works with either.
    """
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        yield from read_classification_xlsx(path, required=required)
    else:
        yield from read_classification_tsv(path, required=required)


def read_classification_tsv(
    path: Path, *, required: Sequence[str] = ("code", "title")
) -> Iterator[tuple[str, list[dict[str, str]]]]:
    """Yield `(code, rows)` for each code in an official classification TSV."""
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t", quotechar='"')
        try:
            header = next(reader)
        except StopIteration:
            raise TerminologyFormatError(f"{path} is empty; expected a header row") from None

        fields = _map_header(header, path, required)
        groups = _group_rows(fields, enumerate(reader, start=2), path)

    yield from groups.items()


# How far into a sheet to look for the header row. The published workbooks put
# it first, but a title or a validity date above it would be unremarkable.
MAX_HEADER_SCAN_ROWS = 25


def read_classification_xlsx(
    path: Path, *, required: Sequence[str] = ("code", "title")
) -> Iterator[tuple[str, list[dict[str, str]]]]:
    """Yield `(code, rows)` for each code in an official classification workbook.

    The published workbooks carry a human-facing "Läs mig" sheet alongside the
    data, so the header row is located by search rather than assumed: the first
    sheet with a row mapping to every required column wins. That also tolerates
    a title row above the header.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - declared dependency
        raise TerminologyFormatError(
            "reading .xlsx classification files requires openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        seen_headers: list[str] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            buffered: list[tuple[int, Sequence[object]]] = []
            fields: list[str | None] | None = None

            for line_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if fields is None:
                    if line_no > MAX_HEADER_SCAN_ROWS:
                        break
                    try:
                        fields = _map_header(row, path, required)
                    except TerminologyFormatError:
                        continue
                    seen_headers.append(sheet_name)
                    continue
                buffered.append((line_no, row))

            if fields is not None:
                logger.info(
                    "classification_workbook_sheet_selected",
                    extra={"path": str(path), "sheet": sheet_name, "rows": len(buffered)},
                )
                if "parent" not in fields:
                    # Not fatal, but it silently flattens the hierarchy, and a
                    # flat hierarchy is easy to miss until `chapter` is empty
                    # everywhere.
                    logger.warning(
                        "classification_source_has_no_parent_column",
                        extra={"path": str(path), "sheet": sheet_name},
                    )
                yield from _group_rows(fields, buffered, path).items()
                return

        raise TerminologyFormatError(
            f"{path}: no sheet has a header row with the required column(s) "
            f"{list(required)} within the first {MAX_HEADER_SCAN_ROWS} rows. "
            f"Sheets examined: {workbook.sheetnames!r}."
        )
    finally:
        workbook.close()


def collect_synonyms(
    rows: Sequence[Mapping[str, str]], *, extra_fields: Sequence[str]
) -> list[str]:
    """Gather alternative surface forms for a code, order-preserving and deduped.

    `Utesluter` is deliberately NOT a synonym source. It lists what the code
    does *not* cover ("Gammal hjartinfarkt (I25.2)" under I21), so indexing it
    would make a code retrievable by the very terms that rule it out. The same
    reasoning excludes Anmarkning and Kodningsinformation, which are guidance
    prose rather than names for the concept.
    """
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        for field in extra_fields:
            value = row.get(field, "").strip()
            if value and value.casefold() not in seen:
                seen.add(value.casefold())
                out.append(value)
    return out


def assign_hierarchy(concepts: list[Concept]) -> list[Concept]:
    """Fill in `is_leaf` and `chapter` from the parent links in the same load.

    `is_leaf` is derived structurally -- a concept is a leaf when no other
    concept in the load names it as parent -- rather than read from the
    publisher's Kodniva column, whose value set is not documented in the file
    description PDFs.

    `chapter` is the topmost ancestor reached by walking parent links, which for
    ICD-10-SE is the chapter interval (for example `I00-I99`).
    """
    by_code = {concept.code: concept for concept in concepts}
    parents = {concept.parent_code for concept in concepts if concept.parent_code}

    for concept in concepts:
        concept.is_leaf = concept.code not in parents
        concept.chapter = _root_of(concept, by_code)
    return concepts


def _root_of(concept: Concept, by_code: Mapping[str, Concept]) -> str | None:
    seen: set[str] = {concept.code}
    current = concept
    while current.parent_code and current.parent_code in by_code:
        if current.parent_code in seen:  # defensive: a cycle in the source file
            break
        seen.add(current.parent_code)
        current = by_code[current.parent_code]
    return current.code if current.code != concept.code else None


def build_search_text(concept: Concept) -> str:
    """The names of a concept: preferred term plus synonyms.

    This is what trigram similarity compares against. Descriptions are
    deliberately excluded here: trigram matching exists to survive a misspelled
    *name*, and running it over prose produces noise rather than tolerance.
    """
    return " ".join([concept.preferred_term, *concept.synonyms]).strip()


def build_synonym_text(concept: Concept) -> str:
    return " ".join(concept.synonyms).strip()
